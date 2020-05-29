#CODE BY YASHARTH DUBEY
#IIIT DHARWAD
import tkinter
import xlrd
import xlwt
import openpyxl as op
import uuid
import smtplib
import os
import time
import sys
import xlsxwriter
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders
from win32_setctime import setctime
try:
    def registerUser():
        window.destroy()
        window33 = tkinter.Tk()
        window33.geometry("450x120")
        window33.title("Crop Management System")
        tkinter.Label(window33,text = "PLEASE PROVIDE THE FIELDS").pack()
        q1 = tkinter.Label(window33,text = "USERNAME:")
        q1.place(x = 40, y = 20)
        q2 = tkinter.Label(window33,text = "PASSWORD:")
        q2.place(x = 40, y = 50)
        w111 = tkinter.Entry(window33,bd = 3,width = 40)
        w111.place(x= 150 , y = 20)
        w211 = tkinter.Entry(window33,bd = 3,width = 40)
        w211.place(x= 150 , y = 50)
        loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/USERID.xlsx")
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)
        def writea():
            flag = 0
            z1 = w111.get()
            z2 = w211.get()
            for i in range(0,sheet.nrows):
                    if(sheet.cell_value(i,0)==z1):
                        tkinter.messagebox.showinfo(title = "ERROR", message = "USERID TAKEN")
                        flag = 1
            if(z1 == "") or (z2 == ""):
                tkinter.messagebox.showinfo(title="EMPTY SPACE",message="Please Fill the Details")
            elif (flag == 0):
                wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/USERID.xlsx")
                ws = wb.get_sheet_by_name("Sheet1")
                ws.append([z1,int(z2)])
                wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/USERID.xlsx')
                wb.close()
                tkinter.messagebox.showinfo(title = "FILE SAVED", message = "Your Data is saved")
                tkinter.messagebox.showinfo(title="THANKS",message="THANKS FOR USING OUR GUI")
                window33.destroy()
        bt90 = tkinter.Button(window33,text = "ENTER",command = writea)
        bt90.place(x = 200,y = 80 )
        window33.mainloop()
    def loginUser():
        window.destroy()
        window41 = tkinter.Tk()
        window41.geometry("400x120")
        window41.title("Fund Tracking System")
        window41.resizable(0,0)
        label2 = tkinter.Label(window41,text = "LOGIN ID:")
        label2.place(x = 100,y = 40)
        et1 = tkinter.Entry(window41, bd = 3)
        et1.place(x = 160 , y = 40)
        def login():
            flag = 0
            loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/USERID.xlsx")
            wb = xlrd.open_workbook(loc)
            sheet = wb.sheet_by_index(0)
            for i in range(0,sheet.nrows):
                if(sheet.cell_value(i,0)==et1.get()):
                    flag = 1
                    window41.destroy()
                    window3 = tkinter.Tk()
                    window3.title("Fund Tracking System")
                    window3.resizable(0,0)
                    window3.geometry("400x120")
                    label3 = tkinter.Label(window3,text = "PASSWORD:")
                    label3.place(x = 90,y = 40)
                    et2 = tkinter.Entry(window3, bd = 3, show = "*")
                    et2.place(x = 160 , y = 40)
                    def password():
                        flag1 = 0
                        for j in range(0,sheet.nrows):
                            if(sheet.cell_value(j,1)==int(et2.get())):
                                flag1 = 1
                        if(flag1 == 0):
                            tkinter.messagebox.showerror(title = "WRONG INPUT", message = "Program Terminating!")
                            window3.destroy()
                            exit()
                        if(flag1 == 1):
                            window3.destroy()
                            window4 = tkinter.Tk()
                            window4.geometry("400x150")
                            window4.title("Fund Tracking System")
                            window4.resizable(0,0)
                            tkinter.Label(window4,text = "WELCOME USER").pack()
                            def addData():
                                window4.destroy()
                                window10 = tkinter.Tk()
                                window10.geometry("450x220")
                                window10.title("Fund Tracking System")
                                tkinter.Label(window10,text = "PLEASE PROVIDE THE FIELDS").pack()
                                q1 = tkinter.Label(window10,text = "USER NAME:")
                                q1.place(x = 40, y = 20)
                                q2 = tkinter.Label(window10,text = "AMOUNT:")
                                q2.place(x = 40, y = 50)
                                q3 = tkinter.Label(window10,text = "EMAIL:")
                                q3.place(x = 40, y = 80)
                                q4 = tkinter.Label(window10,text = "DATE:")
                                q4.place(x = 40, y = 110)
                                q5 = tkinter.Label(window10,text = "PHONE NUMBER:")
                                q5.place(x = 40, y = 140)
                                w1 = tkinter.Entry(window10,bd = 3,width = 40)
                                w1.place(x= 150 , y = 20)
                                w2 = tkinter.Entry(window10,bd = 3,width = 40)
                                w2.place(x= 150 , y = 50)
                                w3 = tkinter.Entry(window10,bd = 3,width = 40)
                                w3.place(x= 150 , y = 80)
                                w4 = tkinter.Entry(window10,bd = 3,width = 40)
                                w4.place(x= 150 , y = 110)
                                w5 = tkinter.Entry(window10,bd = 3,width = 40)
                                w5.place(x= 150 , y = 140)
                                def write():
                                    z1 = w1.get()
                                    z2 = w2.get()
                                    z3 = w3.get()
                                    z4 = w4.get()
                                    z5 = w5.get()
                                    if (z1 == "") or (z2 == "") or (z3 == "") or (z4 == "") or (z5 == ""):
                                        tkinter.messagebox.showinfo(title="EMPTY SPACE",message="Please Fill the Details")
                                    else:
                                        wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx")
                                        ws = wb.get_sheet_by_name("Sheet1")
                                        c = str(uuid.uuid1())
                                        ws.append([z1,c,float(z2),z3,z4,int(z5),"ONLINE"])
                                        wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx')
                                        wb.close()
                                        s = smtplib.SMTP('smtp.gmail.com', 587) 
                                        s.starttls() 
                                        s.login("**********", "***********")
                                        s.sendmail("************",z3,"Your transaction is succesful and the id will be sent to you,Thanks for deposit")
                                        s.sendmail("***********",z3, c) 
                                        s.quit()
                                        tkinter.messagebox.showinfo(title = "FILE SAVED", message = "Thanks for deposit")
                                        window10.destroy()
                                bt90 = tkinter.Button(window10,text = "ENTER",command = write)
                                bt90.place(x = 200,y = 180 )
                            bt7 = tkinter.Button(window4, text = "DEPOSIT FUND",command = addData)
                            bt7.place(x = 150 , y=40) 
                            def requestU():
                                window4.destroy()
                                window18 = tkinter.Tk()
                                window18.geometry("400x160")
                                window18.title("Fund Tracking System")
                                window18.resizable(0,0)
                                q1 = tkinter.Label(window18,text = "USER NAME:")
                                q1.place(x = 40, y = 20)
                                q2 = tkinter.Label(window18,text = "EMAIL:")
                                q2.place(x = 40, y = 50)
                                w1 = tkinter.Entry(window18,bd = 3,width = 40)
                                w1.place(x= 120 , y = 20)
                                w2 = tkinter.Entry(window18,bd = 3,width = 40)
                                w2.place(x= 120 , y = 50)
                                def write():
                                    z1 = w1.get()
                                    z2 = w2.get()
                                    if (z1 == "") or (z2 == ""):
                                        tkinter.messagebox.showinfo(title="EMPTY SPACE",message="Please Fill the Details")
                                    else:
                                        wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
                                        ws = wb.get_sheet_by_name("Sheet1")
                                        ws.append([z1,z2])
                                        wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx')
                                        wb.close()
                                        tkinter.messagebox.showinfo(title = "APPROVED", message = "Your Request Taken")
                                        window18.destroy()
                                bt90 = tkinter.Button(window18,text = "ENTER",command = write)
                                bt90.place(x = 200,y = 100 )
                                window18.mainloop()
                            bt45 = tkinter.Button(window4,text = "REQUEST YOUR DATA",command = requestU)
                            bt45.place(x = 130 , y = 80)
                            window4.mainloop()
                    bt3 = tkinter.Button(window3, text = "ENTER",command = password)
                    bt3.place(x = 160, y = 80)
                    window3.mainloop()
            if flag == 0:
                tkinter.messagebox.showerror(title = "WRONG INPUT", message = "Program Terminating!")
                window2.destroy()
                exit()
        bt2 = tkinter.Button(window41, text = "LOGIN",command = login)
        bt2.place(x = 160, y = 80)       
        window41.mainloop()
    def loginAdmin():
        window.destroy()
        window2 = tkinter.Tk()
        window2.title("Fund Tracking System")
        window2.resizable(0,0)
        window2.geometry("400x120")
        label2 = tkinter.Label(window2,text = "LOGIN ID:")
        label2.place(x = 100,y = 40)
        et1 = tkinter.Entry(window2, bd = 3)
        et1.place(x = 160 , y = 40)
        def login():
            flag = 0
            loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/LOGINID.xlsx")
            wb = xlrd.open_workbook(loc)
            sheet = wb.sheet_by_index(0)
            for i in range(0,sheet.nrows):
                if(sheet.cell_value(i,0)==et1.get()):
                    flag = 1
                    window2.destroy()
                    window3 = tkinter.Tk()
                    window3.title("Fund Tracking System")
                    window3.resizable(0,0)
                    window3.geometry("400x120")
                    label3 = tkinter.Label(window3,text = "PASSWORD:")
                    label3.place(x = 90,y = 40)
                    et2 = tkinter.Entry(window3, bd = 3, show = "*")
                    et2.place(x = 160 , y = 40)
                    def password():
                        flag1 = 0
                        for j in range(0,sheet.nrows):
                            if(sheet.cell_value(j,1)==int(et2.get())):
                                flag1 = 1
                        if(flag1 == 0):
                            tkinter.messagebox.showerror(title = "WRONG INPUT", message = "Program Terminating!")
                            window3.destroy()
                            exit()
                        if(flag1 == 1):
                            window3.destroy()
                            window4 = tkinter.Tk()
                            window4.geometry("400x200")
                            window4.title("Fund Tracking System")
                            window4.resizable(0,0)
                            tkinter.Label(window4,text = "WELCOME ADMIN").pack()
                            def vieway():
                                window4.destroy()
                                window5 = tkinter.Tk()
                                window5.geometry("400x120")
                                window5.title("Fund Tracking System")
                                label5 = tkinter.Label(window5, text = "NAME ON THE USER:")
                                label5.place(x = 40 , y = 40)
                                window5.resizable(0,0)
                                et3 = tkinter.Entry(window5, bd = 3)
                                et3.place(x = 160, y = 40)
                                def showy():
                                    h = str(et3.get())
                                    window5.destroy()
                                    loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx")
                                    wb = xlrd.open_workbook(loc)
                                    sheet = wb.sheet_by_index(0)
                                    window6 = tkinter.Tk()
                                    window6.geometry("400x200")
                                    window6.title("Fund Tracking System")
                                    tkinter.Label(window6, text = "YOUR DATA").pack()
                                    l11 = tkinter.Label(window6, text = "NAME OF USER")
                                    l11.place(x = 5,y=20)
                                    l12 = tkinter.Label(window6, text = "TRANSACTION ID")
                                    l12.place(x = 245,y=20)
                                    l13 = tkinter.Label(window6, text = "AMOUNT")
                                    l13.place(x = 450,y=20)
                                    l14 = tkinter.Label(window6, text = "EMAIL")
                                    l14.place(x = 620,y=20)
                                    l15 = tkinter.Label(window6, text = "DATE")
                                    l15.place(x = 790,y=20)
                                    l16 = tkinter.Label(window6, text = "PHONE NUMBER")
                                    l16.place(x = 880,y=20)
                                    l17 = tkinter.Label(window6,text = "MODE")
                                    l17.place(x = 1000 , y=20)
                                    r = 40
                                    for i in range(0,sheet.nrows):
                                        if(sheet.cell_value(i,0)==h):
                                            l1 = tkinter.Label(window6, text = sheet.cell_value(i,0))
                                            l1.place(x = 2 , y = r)
                                            l2 = tkinter.Label(window6, text = sheet.cell_value(i,1))
                                            l2.place(x = 180 , y = r)
                                            l3 = tkinter.Label(window6, text = str(sheet.cell_value(i,2)))
                                            l3.place(x = 450, y = r)
                                            l4 = tkinter.Label(window6, text = sheet.cell_value(i,3))
                                            l4.place(x = 570, y = r)
                                            l5 = tkinter.Label(window6, text = sheet.cell_value(i,4))
                                            l5.place(x = 750, y = r)
                                            l6 = tkinter.Label(window6, text = str(int(sheet.cell_value(i,5))))
                                            l6.place(x = 875 , y = r)
                                            l7 = tkinter.Label(window6,text = sheet.cell_value(i,6))
                                            l7.place(x = 1000 , y = r)
                                            r = r + 20
                                    window6.mainloop()
                                bt6 = tkinter.Button(window5, text = "ENTER" , command = showy)
                                bt6.place(x = 140 , y = 80)           
                                window5.mainloop()                         
                            bt5 = tkinter.Button(window4,text = "VIEW USER TRANSACTION",command = vieway)
                            bt5.place(x = 125, y = 40)
                            def addData():
                                window4.destroy()
                                window10 = tkinter.Tk()
                                window10.geometry("450x220")
                                window10.title("Fund Tracking System")
                                tkinter.Label(window10,text = "PLEASE PROVIDE THE FIELDS").pack()
                                q1 = tkinter.Label(window10,text = "USER NAME:")
                                q1.place(x = 40, y = 20)
                                q2 = tkinter.Label(window10,text = "AMOUNT:")
                                q2.place(x = 40, y = 50)
                                q3 = tkinter.Label(window10,text = "EMAIL:")
                                q3.place(x = 40, y = 80)
                                q4 = tkinter.Label(window10,text = "DATE:")
                                q4.place(x = 40, y = 110)
                                q5 = tkinter.Label(window10,text = "PHONE NUMBER:")
                                q5.place(x = 40, y = 140)
                                w1 = tkinter.Entry(window10,bd = 3,width = 40)
                                w1.place(x= 150 , y = 20)
                                w2 = tkinter.Entry(window10,bd = 3,width = 40)
                                w2.place(x= 150 , y = 50)
                                w3 = tkinter.Entry(window10,bd = 3,width = 40)
                                w3.place(x= 150 , y = 80)
                                w4 = tkinter.Entry(window10,bd = 3,width = 40)
                                w4.place(x= 150 , y = 110)
                                w5 = tkinter.Entry(window10,bd = 3,width = 40)
                                w5.place(x= 150 , y = 140)
                                def write():
                                    z1 = w1.get()
                                    z2 = w2.get()
                                    z3 = w3.get()
                                    z4 = w4.get()
                                    z5 = w5.get()
                                    if (z1 == "") or (z2 == "") or (z3 == "") or (z4 == "") or (z5 == ""):
                                        tkinter.messagebox.showinfo(title="EMPTY SPACE",message="Please Fill the Details")
                                    else:
                                        wb = op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx")
                                        ws = wb.get_sheet_by_name("Sheet1")
                                        c = str(uuid.uuid1())
                                        ws.append([z1,c,float(z2),z3,z4,int(z5),"OFFLINE"])
                                        wb.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx')
                                        wb.close()
                                        s = smtplib.SMTP('smtp.gmail.com', 587) 
                                        s.starttls() 
                                        s.login("yashdubeywinner@gmail.com", "***********")
                                        s.sendmail("yashdubeywinner@gmail.com",z3, c) 
                                        s.quit()
                                        tkinter.messagebox.showinfo(title = "FILE SAVED", message = "Your Data is saved")
                                        tkinter.messagebox.showinfo(title="THANKS",message="THANKS FOR USING OUR GUI")
                                        window10.destroy()
                                bt90 = tkinter.Button(window10,text = "ENTER",command = write)
                                bt90.place(x = 200,y = 180 )
                            bt10 = tkinter.Button(window4,text = "ADD DATA",command = addData)
                            bt10.place(x = 155, y = 80)
                            def shown():
                                window4.destroy()
                                loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
                                wb = xlrd.open_workbook(loc)
                                sheet = wb.sheet_by_index(0)
                                window5 = tkinter.Tk()
                                window5.geometry("400x120")
                                window5.title("Fund Tracking System")
                                l11 = tkinter.Label(window5, text = "NAME OF USER")
                                l11.place(x = 5,y=20)
                                l12 = tkinter.Label(window5, text = "EMAIL")
                                l12.place(x = 245,y=20)
                                r = 40
                                for i in range(0,sheet.nrows):
                                    l1 = tkinter.Label(window5, text = sheet.cell_value(i,0))
                                    l1.place(x = 2 , y = r)
                                    l2 = tkinter.Label(window5, text = sheet.cell_value(i,1))
                                    l2.place(x = 180 , y = r)
                                    r = r + 20
                                def sendr():
                                    workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx")
                                    workbook.close()
                                    loc = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/funds.xlsx")
                                    wb = xlrd.open_workbook(loc)
                                    sheet = wb.sheet_by_index(0)
                                    wb1 =  op.load_workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx")
                                    ws = wb1.get_sheet_by_name("Sheet1")
                                    loc1 = ("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
                                    wb2 = xlrd.open_workbook(loc1)
                                    sheet2 = wb2.sheet_by_index(0)
                                    window5.destroy()
                                    for i in range(0,sheet2.nrows):
                                        os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx")
                                        workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx")
                                        workbook.close()
                                        for j in range(0,sheet.nrows):
                                            if(sheet2.cell_value(i,1)==sheet.cell_value(j,3)):
                                                ws.append([sheet.cell_value(j,0),sheet.cell_value(j,1),sheet.cell_value(j,2),sheet.cell_value(j,3),sheet.cell_value(j,4),sheet.cell_value(j,5),sheet.cell_value(j,6)])
                                        wb1.save(filename = 'C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx')
                                        wb1.close()
                                        fromaddr = "yashdubeywinner@gmail.com"
                                        toaddr = str(sheet2.cell_value(i,1))
                                        msg = MIMEMultipart()
                                        msg['From'] = fromaddr
                                        msg['To'] = toaddr 
                                        msg['Subject'] = "Reply of the Request Sent"
                                        body = "In return of your request we had sent you a transation file which will show your tranctions of the id requested"
                                        msg.attach(MIMEText(body, 'plain')) 
                                        filename = "Transaction.xlsx"
                                        attachment = open("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Transaction.xlsx", "rb")
                                        p = MIMEBase('application', 'octet-stream')
                                        p.set_payload((attachment).read())
                                        encoders.encode_base64(p)
                                        p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                                        msg.attach(p)
                                        s = smtplib.SMTP('smtp.gmail.com', 587)
                                        s.starttls()
                                        s.login(fromaddr, "**************")
                                        text = msg.as_string() 
                                        s.sendmail(fromaddr, toaddr, text)
                                        s.quit()  
                                    
                                bt45 = tkinter.Button(window5,text = "SEND REPLY",command = sendr)
                                bt45.pack()
                            bt11 = tkinter.Button(window4,text = "SEE REQUEST",command = shown)
                            bt11.place(x = 145, y = 120)
                            window4.mainloop()
                    bt3 = tkinter.Button(window3, text = "ENTER",command = password)
                    bt3.place(x = 160, y = 80)
                    window3.mainloop()
            if flag == 0:
                tkinter.messagebox.showerror(title = "WRONG INPUT", message = "Program Terminating!")
                window2.destroy()
                exit()
        bt2 = tkinter.Button(window2, text = "LOGIN",command = login)
        bt2.place(x = 160, y = 80)
    window = tkinter.Tk()
    window.title("Fund Tracking System")
    window.geometry("400x200")
    tkinter.Label(window,text = "WELCOME TO FUND TRACKING SYSTEM").pack()
    bt = tkinter.Button(window, text = "LOGIN AS ADMIN", command = loginAdmin)
    bt.place(x = 160 , y = 80)
    bt1 = tkinter.Button(window,text = "LOGIN AS USER", command = loginUser)
    bt1.place(x = 165 , y = 120)
    bt0 = tkinter.Button(window,text = "REGISTER AS USER", command = registerUser)
    bt0.place(x=158, y = 40)
    current_time = time.time()
    creation_time = os.path.getctime("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
    if (current_time - creation_time) // (24 * 3600) >= 1:
        os.remove("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
        workbook = xlsxwriter.Workbook("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx")
        workbook.close()
        setctime("C:/Users/yasharth dubey/Documents/Python Scripts/Fund tracking System/Request.xlsx", current_time)
    bt4 = tkinter.Button(window,text = "EXIT", command = exit )
    bt4.place(x = 190 , y = 160)
    window.mainloop()
except:
    tkinter.messagebox.showerror(title = "ERROR", message = "PROGRAM TERMINATING")
finally:
    tkinter.messagebox.showinfo(title = "THANKS", message = "THANKS FOR USING")
